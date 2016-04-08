# -*- coding: utf-8 -*-
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, colors
from django.http import HttpResponse
from django.contrib import admin
from django import forms
from .models import Questionnaire, Question, MailConfig, Survey, SurveyQuestion


def export_survey_report(modeladmin, request, queryset):
    f = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Résulats"
    ws.append([
            'Num Enquête', 'Enquête', 'Titre', 'Nom', 'Prénom', 'Adresse Mail',
            'Complétée', 'Date évènement', 'Date d\'envoi',
            'Date de réception', 'Score enquête', 'Score maximum',
            'Questionnaire', 'Question', 'Récapitulative', 'Score question'
        ])
    for row in ws.iter_rows("A1:P1"):
        for cell in row:
            cell.fill = PatternFill(
                start_color=colors.BLACK,
                end_color=colors.BLACK,
                fill_type='solid')
            cell.font = Font(color=colors.WHITE)
    for survey in queryset:
        for question in survey.surveyquestion_set.all():
            ws.append([
                    survey.id, survey.subject, survey.title, survey.firstname,
                    survey.lastname, survey.email, survey.completed(),
                    survey.event_date, survey.date_sent, survey.date_received,
                    survey.score(), survey.score_max(),
                    survey.questionnaire.title, question.question_text,
                    question.record, question.score
                ])
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column] = max((
                        dims.get(cell.column, 0),
                        len(str(cell.value))
                        ))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value + 2
    wb.save(f)
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=Report.xlsx'
    response.write(f.getvalue())
    f.close()
    return response
export_survey_report.short_description = "Exporter les résultats vers Excel"


class MailConfigForm(forms.ModelForm):
    password = forms.CharField(widget=forms.PasswordInput(render_value=True))

    class Meta:
        model = MailConfig
        fields = ('name', 'domain', 'sender', 'password', 'subject', 'message')


class MailConfigAdmin(admin.ModelAdmin):
    model = MailConfig
    form = MailConfigForm

admin.site.register(MailConfig, MailConfigAdmin)


class QuestionForm(forms.ModelForm):
    question_text = forms.CharField(
        widget=forms.TextInput(attrs={'size': '120'}))

    class Meta:
        model = Question
        fields = ('question_text', 'record')


class QuestionsInline(admin.TabularInline):
    model = Question
    form = QuestionForm
    fields = ['question_text', 'record']
    extra = 0


class QuestionnaireAdmin(admin.ModelAdmin):
    inlines = [QuestionsInline]

admin.site.register(Questionnaire, QuestionnaireAdmin)


class SurveyQuestionForm(forms.ModelForm):
    question_text = forms.CharField(
        widget=forms.TextInput(attrs={'size': '120'}))

    class Meta:
        model = SurveyQuestion
        fields = ('question_text', 'record', 'score')


class SurveyQuestionInline(admin.TabularInline):
    model = SurveyQuestion
    form = SurveyQuestionForm
    fields = ['question_text', 'record', 'score']
#    readonly_fields = ['question_text', 'record']
    can_delete = False
    extra = 0

    def has_add_permission(self, request):
        return False


class SurveyForm(forms.ModelForm):
    class Meta:
        model = Survey
        exclude = ('current_questionnaire',)


class SurveyAdmin(admin.ModelAdmin):
    list_display = [
        'email', 'subject', 'date_sent', 'date_received', 'sent', 'completed',
        'score', 'score_max'
        ]
    form = SurveyForm
    readonly_fields = ['token']
    inlines = [SurveyQuestionInline]
    actions = [export_survey_report]

admin.site.register(Survey, SurveyAdmin)
